from flask import Flask, request, render_template, send_file, jsonify
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER

def highlight_matches(file1_path, file2_path, col1_name, col2_name):
    try:
        # Carica i file Excel
        df1 = pd.read_excel(file1_path)
        df2 = pd.read_excel(file2_path)

        # Verifica che le colonne esistano
        if col1_name not in df1.columns:
            return None, f"Errore: La colonna '{col1_name}' non esiste nel primo file."

        if col2_name not in df2.columns:
            return None, f"Errore: La colonna '{col2_name}' non esiste nel secondo file."

        # Converte i dati in set per il confronto
        cod_set = set(df1[col1_name].astype(str).str.strip())

        # Carica il file con openpyxl per la formattazione
        wb = load_workbook(file2_path)
        ws = wb.active

        # Trova l'indice della colonna corretta
        col_idx = df2.columns.get_loc(col2_name) + 1

        # Definisce il colore giallo per evidenziare
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Scansiona la colonna e applica il colore alle corrispondenze
        match_count = 0
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value and str(cell_value).strip() in cod_set:
                ws.cell(row=row, column=col_idx).fill = yellow_fill
                match_count += 1

        # Se nessun valore corrisponde, avvisa l'utente
        if match_count == 0:
            return None, "Nessuna corrispondenza trovata tra i due file."

        # Salva il nuovo file con le corrispondenze evidenziate
        output_path = os.path.join(UPLOAD_FOLDER, "File_2_Highlighted.xlsx")
        wb.save(output_path)
        wb.close()

        return output_path, None

    except Exception as e:
        return None, f"Errore nel confronto dei file: {str(e)}"

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/upload", methods=["POST"])
def upload_files():
    try:
        if "file1" not in request.files or "file2" not in request.files:
            return jsonify({"error": "Errore: Carica entrambi i file"}), 400

        file1 = request.files["file1"]
        file2 = request.files["file2"]
        col1_name = request.form.get("col1_name", "").strip()
        col2_name = request.form.get("col2_name", "").strip()

        if not file1.filename or not file2.filename or not col1_name or not col2_name:
            return jsonify({"error": "Errore: Carica entrambi i file e specifica le colonne"}), 400

        file1_path = os.path.join(UPLOAD_FOLDER, "File_1.xlsx")
        file2_path = os.path.join(UPLOAD_FOLDER, "File_2.xlsx")

        file1.save(file1_path)
        file2.save(file2_path)

        output_path, error = highlight_matches(file1_path, file2_path, col1_name, col2_name)

        if error:
            return jsonify({"error": error}), 400

        return send_file(output_path, as_attachment=True)

    except Exception as e:
        return jsonify({"error": f"Errore durante l'upload dei file: {str(e)}"}), 500

if __name__ == "__main__":
    app.run(debug=True)